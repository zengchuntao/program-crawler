"""Tests for the new Research Agent Excel exporter."""

from pathlib import Path

from openpyxl import load_workbook

from app.export.xlsx_exporter import ResearchExporter
from contracts.models import (
    AgentLogStep,
    EvidenceItem,
    Finding,
    ResearchGoal,
    ResearchResult,
    ResearchStatus,
    ScreenshotType,
)


def _make_test_result() -> ResearchResult:
    """Create a test ResearchResult with sample data."""
    return ResearchResult(
        query="Find MIT CS GPA and language requirements",
        goals=ResearchGoal(
            target_entities=["MIT CS Master's"],
            fields_requested=["gpa_requirement", "language_requirement"],
        ),
        findings=[
            Finding(
                finding_id="FIND-test0001",
                entity="MIT CS Master's",
                field_name="gpa_requirement",
                value="3.5",
                confidence=0.95,
                source_url="https://mit.edu/admissions",
                evidence_id="EVID-test0001",
                screenshot_file="evidence/agent/FIND-test0001_full.png",
            ),
            Finding(
                finding_id="FIND-test0002",
                entity="MIT CS Master's",
                field_name="language_requirement",
                value="IELTS 7.0 / TOEFL 100",
                confidence=0.90,
                source_url="https://mit.edu/admissions",
                evidence_id="EVID-test0002",
                screenshot_file="evidence/agent/FIND-test0002_full.png",
            ),
        ],
        evidence_items=[
            EvidenceItem(
                evidence_id="EVID-test0001",
                program_id="agent",
                field="gpa_requirement",
                source_url="https://mit.edu/admissions",
                screenshot_files=["evidence/agent/FIND-test0001_full.png"],
                screenshot_type=ScreenshotType.FULL_PAGE,
            ),
            EvidenceItem(
                evidence_id="EVID-test0002",
                program_id="agent",
                field="language_requirement",
                source_url="https://mit.edu/admissions",
                screenshot_files=["evidence/agent/FIND-test0002_full.png"],
                screenshot_type=ScreenshotType.FULL_PAGE,
            ),
        ],
        agent_log=[
            AgentLogStep(
                step=1, action_type="search",
                url_or_query="MIT CS admissions GPA TOEFL",
                reasoning="Starting search",
            ),
            AgentLogStep(
                step=2, action_type="visit",
                url_or_query="https://mit.edu/admissions",
                reasoning="Visit official page", found_info=True,
            ),
            AgentLogStep(
                step=3, action_type="done",
                reasoning="All fields found", found_info=True,
            ),
        ],
        status=ResearchStatus.COMPLETE,
        duration_s=12.5,
        total_steps=3,
        pages_visited=1,
    )


def test_research_exporter_creates_file(tmp_path: Path):
    exporter = ResearchExporter()
    result = _make_test_result()
    exporter.load_result(result)

    out_path = tmp_path / "results.xlsx"
    saved = exporter.save(out_path)
    assert saved.exists()
    assert saved.stat().st_size > 0


def test_research_exporter_has_4_sheets(tmp_path: Path):
    exporter = ResearchExporter()
    exporter.load_result(_make_test_result())

    out_path = tmp_path / "results.xlsx"
    exporter.save(out_path)

    wb = load_workbook(str(out_path))
    sheet_names = wb.sheetnames
    assert "Findings" in sheet_names
    assert "Summary" in sheet_names
    assert "Evidence" in sheet_names
    assert "Agent Log" in sheet_names


def test_findings_sheet_data(tmp_path: Path):
    exporter = ResearchExporter()
    exporter.load_result(_make_test_result())

    out_path = tmp_path / "results.xlsx"
    exporter.save(out_path)

    wb = load_workbook(str(out_path))
    fs = wb["Findings"]
    # Header row + 2 data rows
    assert fs.max_row == 3
    # Check header
    assert fs.cell(row=1, column=1).value == "Finding ID"
    # Check data
    assert fs.cell(row=2, column=1).value == "FIND-test0001"
    assert fs.cell(row=2, column=4).value == "3.5"


def test_summary_sheet_data(tmp_path: Path):
    exporter = ResearchExporter()
    exporter.load_result(_make_test_result())

    out_path = tmp_path / "results.xlsx"
    exporter.save(out_path)

    wb = load_workbook(str(out_path))
    ss = wb["Summary"]
    # Header + 1 entity row
    assert ss.max_row == 2
    assert ss.cell(row=2, column=1).value == "MIT CS Master's"


def test_agent_log_sheet_data(tmp_path: Path):
    exporter = ResearchExporter()
    exporter.load_result(_make_test_result())

    out_path = tmp_path / "results.xlsx"
    exporter.save(out_path)

    wb = load_workbook(str(out_path))
    als = wb["Agent Log"]
    # Header + 3 log steps
    assert als.max_row == 4
    assert als.cell(row=2, column=2).value == "search"
