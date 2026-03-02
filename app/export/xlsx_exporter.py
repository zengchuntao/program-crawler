"""Excel exporter — writes results.xlsx with 4 sheets (Research Agent) or legacy 2 sheets."""

from __future__ import annotations

from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font

from contracts.models import (
    AgentLogStep,
    EvidenceItem,
    Finding,
    ProgramRecord,
    ResearchResult,
)


def _load_columns() -> dict:
    """Load column definitions from contracts/columns.yaml."""
    yaml_path = Path(__file__).resolve().parent.parent.parent / "contracts" / "columns.yaml"
    with open(yaml_path, encoding="utf-8") as f:
        return yaml.safe_load(f)


# ---------------------------------------------------------------------------
# New 4-sheet exporter for Research Agent
# ---------------------------------------------------------------------------

class ResearchExporter:
    """Builds an Excel workbook with 4 sheets: Findings, Summary, Evidence, Agent Log."""

    def __init__(self):
        self._columns = _load_columns()
        self._findings: list[Finding] = []
        self._evidence: list[EvidenceItem] = []
        self._agent_log: list[AgentLogStep] = []
        self._entities: set[str] = set()

    def load_result(self, result: ResearchResult) -> None:
        """Load all data from a ResearchResult."""
        self._findings.extend(result.findings)
        self._evidence.extend(result.evidence_items)
        self._agent_log.extend(result.agent_log)
        if result.goals:
            self._entities.update(result.goals.target_entities)
        for f in result.findings:
            self._entities.add(f.entity)

    def save(self, path: str | Path) -> Path:
        """Write the workbook to disk and return the path."""
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        cols = self._columns

        # --- Sheet 1: Findings ---
        self._write_findings_sheet(wb, cols)

        # --- Sheet 2: Summary ---
        self._write_summary_sheet(wb, cols)

        # --- Sheet 3: Evidence ---
        self._write_evidence_sheet(wb, cols)

        # --- Sheet 4: Agent Log ---
        self._write_agent_log_sheet(wb, cols)

        wb.save(str(path))
        return path

    def _write_findings_sheet(self, wb: Workbook, cols: dict) -> None:
        """Write Findings sheet — one row per finding."""
        fs = wb.active
        fs.title = cols["findings_sheet"]["name"]
        sheet_cols = cols["findings_sheet"]["columns"]

        # Header
        for i, col_def in enumerate(sheet_cols, start=1):
            cell = fs.cell(row=1, column=i, value=col_def["header"])
            cell.font = Font(bold=True)
            fs.column_dimensions[cell.column_letter].width = col_def.get("width", 15)

        # Data
        for row_idx, finding in enumerate(self._findings, start=2):
            for col_idx, col_def in enumerate(sheet_cols, start=1):
                key = col_def["key"]
                val = getattr(finding, key, "")
                if val is None:
                    val = ""
                fs.cell(row=row_idx, column=col_idx, value=str(val))

    def _write_summary_sheet(self, wb: Workbook, cols: dict) -> None:
        """Write Summary sheet — one row per entity with all fields."""
        ss = wb.create_sheet(title=cols["summary_sheet"]["name"])
        sheet_cols = cols["summary_sheet"]["columns"]

        # Header
        for i, col_def in enumerate(sheet_cols, start=1):
            cell = ss.cell(row=1, column=i, value=col_def["header"])
            cell.font = Font(bold=True)
            ss.column_dimensions[cell.column_letter].width = col_def.get("width", 15)

        # Build summary data: one row per entity
        entity_data = {}
        for finding in self._findings:
            if finding.entity not in entity_data:
                entity_data[finding.entity] = {"entity": finding.entity}
            entity_data[finding.entity][finding.field_name] = finding.value

        # Determine missing fields per entity
        skip_keys = ("entity", "missing_fields", "status")
        all_fields = {
            c["key"] for c in sheet_cols if c["key"] not in skip_keys
        }
        for entity, data in entity_data.items():
            found_fields = {k for k in data if k != "entity"}
            missing = all_fields - found_fields
            data["missing_fields"] = "; ".join(sorted(missing)) if missing else ""
            data["status"] = "complete" if not missing else "partial"

        # Write rows
        for row_idx, (entity, data) in enumerate(entity_data.items(), start=2):
            for col_idx, col_def in enumerate(sheet_cols, start=1):
                key = col_def["key"]
                val = data.get(key, "")
                if val is None:
                    val = ""
                ss.cell(row=row_idx, column=col_idx, value=str(val))

    def _write_evidence_sheet(self, wb: Workbook, cols: dict) -> None:
        """Write Evidence sheet — evidence index."""
        es = wb.create_sheet(title=cols["evidence_sheet"]["name"])
        sheet_cols = cols["evidence_sheet"]["columns"]

        # Header
        for i, col_def in enumerate(sheet_cols, start=1):
            cell = es.cell(row=1, column=i, value=col_def["header"])
            cell.font = Font(bold=True)
            es.column_dimensions[cell.column_letter].width = col_def.get("width", 15)

        # Build evidence rows linking to findings
        finding_map = {f.evidence_id: f for f in self._findings if f.evidence_id}
        for row_idx, ev in enumerate(self._evidence, start=2):
            finding = finding_map.get(ev.evidence_id)
            row_data = {
                "evidence_id": ev.evidence_id,
                "finding_id": finding.finding_id if finding else "",
                "field": ev.field,
                "screenshot_file": ";".join(ev.screenshot_files) if ev.screenshot_files else "",
                "screenshot_type": ev.screenshot_type.value if ev.screenshot_type else "",
                "source_url": ev.source_url,
            }
            for col_idx, col_def in enumerate(sheet_cols, start=1):
                val = row_data.get(col_def["key"], "")
                es.cell(row=row_idx, column=col_idx, value=str(val))

    def _write_agent_log_sheet(self, wb: Workbook, cols: dict) -> None:
        """Write Agent Log sheet — debug/audit trail."""
        als = wb.create_sheet(title=cols["agent_log_sheet"]["name"])
        sheet_cols = cols["agent_log_sheet"]["columns"]

        # Header
        for i, col_def in enumerate(sheet_cols, start=1):
            cell = als.cell(row=1, column=i, value=col_def["header"])
            cell.font = Font(bold=True)
            als.column_dimensions[cell.column_letter].width = col_def.get("width", 15)

        # Data
        for row_idx, step in enumerate(self._agent_log, start=2):
            for col_idx, col_def in enumerate(sheet_cols, start=1):
                key = col_def["key"]
                val = getattr(step, key, "")
                if val is None:
                    val = ""
                als.cell(row=row_idx, column=col_idx, value=str(val))


# ---------------------------------------------------------------------------
# Legacy 2-sheet exporter (preserved for backward compatibility)
# ---------------------------------------------------------------------------

class XlsxExporter:
    """Builds an Excel workbook with Programs and Evidence sheets (legacy)."""

    def __init__(self):
        self._columns = _load_columns()
        self._programs: list[ProgramRecord] = []
        self._evidence: list[EvidenceItem] = []

    def add_program(self, record: ProgramRecord) -> None:
        self._programs.append(record)

    def add_evidence(self, items: list[EvidenceItem]) -> None:
        self._evidence.extend(items)

    def _get_cell_value(self, record: ProgramRecord, key: str) -> str:
        """Resolve a column key to its string value from a ProgramRecord."""
        # Direct top-level fields
        if hasattr(record, key):
            val = getattr(record, key)
        # Fields inside ExtractedFields
        elif hasattr(record.fields, key):
            val = getattr(record.fields, key)
        # Evidence refs
        elif hasattr(record.evidence, key):
            val = getattr(record.evidence, key)
        else:
            return ""

        if val is None:
            return ""
        if isinstance(val, list):
            return ";".join(str(v) for v in val)
        return str(val)

    def save(self, path: str | Path) -> Path:
        """Write the workbook to disk and return the path."""
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        cols = self._columns

        # --- Programs sheet ---
        ps = wb.active
        ps.title = cols["programs_sheet"]["name"]
        prog_cols = cols["programs_sheet"]["columns"]

        # Header row
        for i, col_def in enumerate(prog_cols, start=1):
            cell = ps.cell(row=1, column=i, value=col_def["header"])
            cell.font = Font(bold=True)
            ps.column_dimensions[cell.column_letter].width = col_def.get("width", 15)

        # Data rows
        for row_idx, record in enumerate(self._programs, start=2):
            for col_idx, col_def in enumerate(prog_cols, start=1):
                val = self._get_cell_value(record, col_def["key"])
                ps.cell(row=row_idx, column=col_idx, value=val)

        # --- Evidence sheet ---
        es = wb.create_sheet(title=cols["evidence_sheet"]["name"])
        ev_cols = cols["evidence_sheet"]["columns"]

        for i, col_def in enumerate(ev_cols, start=1):
            cell = es.cell(row=1, column=i, value=col_def["header"])
            cell.font = Font(bold=True)
            es.column_dimensions[cell.column_letter].width = col_def.get("width", 15)

        for row_idx, item in enumerate(self._evidence, start=2):
            for col_idx, col_def in enumerate(ev_cols, start=1):
                key = col_def["key"]
                val = getattr(item, key, "")
                if val is None:
                    val = ""
                if isinstance(val, list):
                    val = ";".join(str(v) for v in val)
                es.cell(row=row_idx, column=col_idx, value=str(val))

        wb.save(str(path))
        return path
